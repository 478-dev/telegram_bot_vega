import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import DB_PATH

def get_conn():
    return sqlite3.connect(DB_PATH, check_same_thread=False)

def export_to_xlsx(filepath):
    conn = get_conn()
    cur = conn.cursor()

    # Получаем оплаченные заказы с информацией о позициях и категориях
    cur.execute("""
        SELECT 
            o.id,
            c.emoji as cat_emoji,
            c.text as cat_text,
            p.emoji as pos_emoji,
            p.text as pos_text,
            p.price,
            p.category_id,
            o.user_name,
            o.user_info,
            o.payment
        FROM ORDERSS o
        LEFT JOIN POSITIONS p ON o.position_id = p.id
        LEFT JOIN CATEGORIES c ON p.category_id = c.id AND c.id != 0
        WHERE o.payment = 1
        ORDER BY o.id
    """)
    paid_orders = cur.fetchall()

    # Получаем неоплаченные заказы
    cur.execute("""
        SELECT 
            o.id,
            c.emoji as cat_emoji,
            c.text as cat_text,
            p.emoji as pos_emoji,
            p.text as pos_text,
            p.price,
            p.category_id,
            o.user_name,
            o.user_info,
            o.payment
        FROM ORDERSS o
        LEFT JOIN POSITIONS p ON o.position_id = p.id
        LEFT JOIN CATEGORIES c ON p.category_id = c.id AND c.id != 0
        WHERE o.payment = 0
        ORDER BY o.id
    """)
    unpaid_orders = cur.fetchall()

    conn.close()

    # Создаем Excel файл
    wb = Workbook()
    wb.remove(wb.active)

    # Лист с оплаченными
    ws_paid = wb.create_sheet("Оплаченные")
    ws_paid.append(["ID заказа", "Категория", "Позиция", "Цена", "Покупатель", "Доп. инфо", "Статус"])

    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws_paid[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for order in paid_orders:
        order_id, cat_emoji, cat_text, pos_emoji, pos_text, price, category_id, user_name, user_info, payment = order

        # Формируем название категории с эмодзи
        # Если category_id = 0 (корень), оставляем пустую ячейку
        if category_id == 0:
            category_name = ""
        elif cat_emoji and cat_text:
            category_name = f"{cat_emoji} {cat_text}"
        elif cat_text:
            category_name = cat_text
        elif cat_emoji:
            category_name = cat_emoji
        else:
            category_name = ""

        # Формируем название позиции с эмодзи
        if pos_emoji and pos_text:
            position_name = f"{pos_emoji} {pos_text}"
        elif pos_text:
            position_name = pos_text
        elif pos_emoji:
            position_name = pos_emoji
        else:
            position_name = "—"

        ws_paid.append([
            order_id,
            category_name,
            position_name,
            f"{price}₽" if price else "—",
            user_name or "—",
            user_info or "—",
            "✅ Оплачено"
        ])

    # Лист с неоплаченными
    ws_unpaid = wb.create_sheet("Неоплаченные")
    ws_unpaid.append(["ID заказа", "Категория", "Позиция", "Цена", "Покупатель", "Доп. инфо", "Статус"])

    header_fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")

    for cell in ws_unpaid[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for order in unpaid_orders:
        order_id, cat_emoji, cat_text, pos_emoji, pos_text, price, category_id, user_name, user_info, payment = order

        # Формируем название категории с эмодзи
        # Если category_id = 0 (корень), оставляем пустую ячейку
        if category_id == 0:
            category_name = ""
        elif cat_emoji and cat_text:
            category_name = f"{cat_emoji} {cat_text}"
        elif cat_text:
            category_name = cat_text
        elif cat_emoji:
            category_name = cat_emoji
        else:
            category_name = ""

        # Формируем название позиции с эмодзи
        if pos_emoji and pos_text:
            position_name = f"{pos_emoji} {pos_text}"
        elif pos_text:
            position_name = pos_text
        elif pos_emoji:
            position_name = pos_emoji
        else:
            position_name = "—"

        ws_unpaid.append([
            order_id,
            category_name,
            position_name,
            f"{price}₽" if price else "—",
            user_name or "—",
            user_info or "—",
            "❌ Не оплачено"
        ])

    # Установим ширину колонок
    for ws in [ws_paid, ws_unpaid]:
        ws.column_dimensions['A'].width = 12  # ID заказа
        ws.column_dimensions['B'].width = 25  # Категория
        ws.column_dimensions['C'].width = 30  # Позиция
        ws.column_dimensions['D'].width = 12  # Цена
        ws.column_dimensions['E'].width = 25  # Покупатель
        ws.column_dimensions['F'].width = 20  # Доп. инфо
        ws.column_dimensions['G'].width = 15  # Статус

    wb.save(filepath)
    return filepath